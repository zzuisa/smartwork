# -*- encoding: utf-8 -*-
'''
@文件        :excel_util.py
@说明        :
@时间        :2022/09/15 10:26:29
@作者        :awx1192780
@版本        :1.0
'''


from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, Border, Side
from tqdm import tqdm
from util.common_tools import print_pro
from base import constants
import os
import copy
import time


__all__ = ['open_file', 'open_or_add_sheet', 'data_excel', 'export_to_excel']


# 批量插入支持传入'A1' 和 [row,col]形式的数据
# 批量插入数据
# mode 0 横向 1 纵向


def batch_insert(ws, _list, pos=[1, 1], spec_column=None, mode=0, _vertical='center'):
    row, column = 1, 1
    if isinstance(pos, str):
        row, column = [ord(letter)-65+1, int(pos[1:])]
    if isinstance(pos, list):
        row, column = pos
    if mode == 0:
        for index, value in enumerate(_list):
            _target_cell = ws.cell(row=row, column=column+index,
                                   value=value)
            _target_cell.alignment = Alignment(
                horizontal='left', vertical='top', wrapText=True)
            thin = Side(border_style="thin", color="000000")
            double = Side(border_style="double", color="000000")
            _target_cell.border = Border(
                top=double, left=thin, right=thin, bottom=double)
            # 微调格式

            if row == 1:
                ws.cell(row=row, column=column+index).font = Font(bold=True)

            # 居中处理
            # wrapText=True 自动换行

    else:
        for index, value in enumerate(_list):
            ws.cell(row=row+index, column=column,
                    value=value).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    return ws


def batch_insert_report(ws, _list, pos=[1, 1], spec_column=None, mode=0, _vertical='center'):
    row, column = 1, 1
    if isinstance(pos, str):
        row, column = [ord(letter)-65+1, int(pos[1:])]
    if isinstance(pos, list):
        row, column = pos
    smart_width_and_height(ws, row, _list, spec_column)
    if spec_column != None:
        ws.column_dimensions[spec_column].width = 85
        ws.column_dimensions[chrstep(spec_column, -1)].width = 35
    if mode == 0:
        for index, value in enumerate(_list):
            _target_cell = ws.cell(row=row, column=column+index,
                                   value=value)
            _target_cell.alignment = Alignment(
                horizontal='center', vertical='center', wrapText=True)
            thin = Side(border_style="thin", color="000000")
            double = Side(border_style="double", color="000000")
            _target_cell.border = Border(
                top=double, left=thin, right=thin, bottom=double)
            # 微调格式
            if row == 1:
                ws.cell(row=row, column=column+index).font = Font(bold=True)

            # 以下是对 spec_column 列做特殊格式处理
            if spec_column != None:
                if index == ord(spec_column)-65:
                    ws.cell(row=row, column=ord(spec_column)-65+1).alignment = Alignment(
                        horizontal='left', vertical='top', wrapText=True)
        ws.cell(row=1, column=ord(spec_column)-65+1).alignment = Alignment(
            horizontal='center', vertical='center', wrapText=True)

        # 居中处理
        # wrapText=True 自动换行

    else:
        for index, value in enumerate(_list):
            ws.cell(row=row+index, column=column,
                    value=value).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    return ws


# chrstep
def chrstep(c, num):
    return chr(ord(c)+num)


def chr2index(c):
    return ord(c)-65


# 智能调整单元格宽度和高度


def smart_width_and_height(ws, _row, _list: list, spec_column: None):
    for _index, _ in enumerate(_list):
        ws.column_dimensions[chr(_index+65).upper()].width = 12.5 if ws.column_dimensions[chr(
            _index+65).upper()].width < 12.5 else ws.column_dimensions[chr(_index+65).upper()].width
        if len(str(_)) > 30:
            ws.column_dimensions[chr(_index+65).upper()
                                 ].width = 22.5 + (len(str(_))/9) * 3

    # 根据content和desc中的内容调整单元格高度
    tar_height = 24 + 2 * (int(len(_list[chr2index(spec_column)])/9)+1) + \
        ''.join(map(str, _list)).count('\n') * 12

    ws.row_dimensions[_row].height = tar_height
    ws.row_dimensions[1].height = 30

# 设置过滤器和排序字段（有未解决bug，暂不可用）


def set_filter_and_sort(ws, _all_list: list):
    _row = len(_all_list)
    _column = len(_all_list[0])
    # _dates = set([i[2] for i in _all_list])
    # _times = set([i[3] for i in _all_list])
    # print(_row,_column,_dates,_times)
    for _index, _ in enumerate(_all_list):
        ws.column_dimensions[chr(_index+65).upper()].width = 25
    if _index not in [0, 4]:
        ws.auto_filter.add_filter_column(_index, _all_list[_index])
    ws.auto_filter.ref = "A1:{}1".format(chr(_column+65-1).upper())
    ws.auto_filter.add_sort_condition("C2:D{}".format(_row))


def open_file(excel_name_path):
    return Workbook() if not os.path.isfile(excel_name_path) else load_workbook(str(excel_name_path))

# 检查是否存在当前sheet_name, 若不存在则新建并以此命名第一个sheet


def open_or_add_sheet(workbook, sheet_name, template_path=None, template_sheet_name=None):
    worksheet = workbook.active
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    elif len(workbook.sheetnames) != 0 and '月' in workbook.sheetnames[0]:
        worksheet = workbook.create_sheet(sheet_name)
    else:
        if template_path and template_sheet_name not in workbook.sheetnames:
            cp_sheet = copy_sheet_from(template_path, template_sheet_name)
            cp_sheet._parent = workbook
            workbook._add_sheet(cp_sheet)
        worksheet = workbook.create_sheet(sheet_name)
        if workbook.sheetnames[0] != template_sheet_name:
            del workbook[workbook.sheetnames[0]]
    return worksheet

# 执行插入数据


def copy_sheet_from(path, sheet_name):
    wb = open_file(path)
    if sheet_name in wb.sheetnames:
        worksheet = wb[sheet_name]
        worksheet.title = sheet_name
        return worksheet
    else:
        print_pro('不存在此工作簿!', constants.ERROR_PRINT)


# 最终处理，导出excel

def data_excel(excel_name_path, workbook, worksheet, title, data, data_type="count", template_path=None, template_sheet_name=None, spec_column=None):
    if template_sheet_name and template_sheet_name not in workbook.sheetnames:
        cp_sheet = copy_sheet_from(template_path, template_sheet_name)
        cp_sheet._parent = workbook
        workbook._add_sheet(cp_sheet)
    datain = copy.copy(data)  # 浅拷贝
    if isinstance(datain, dict):
        _, values = zip(*datain.items())
    if isinstance(datain, list):
        values = datain
    value_normal = values
    # 如果data_type 是 'count' 则表示计数——统计本月各类事项处理个数，生成统计表
    # 反之 (data_type 是'report') 则表示导出具体交付件，和详情，生成报告表。
    worksheet = batch_insert(worksheet, title, [1, 1], spec_column) if data_type == 'count' else batch_insert_report(
        worksheet, title, [1, 1], spec_column)
    for index, value in enumerate(value_normal):
        worksheet = batch_insert(worksheet, value, [
                                 2+index, 1], spec_column) if data_type == 'count' else batch_insert_report(worksheet, value, [2+index, 1], spec_column)

    # 暂无法使用filter和sort
    # set_filter_and_sort(worksheet, value_normal)
    workbook.save(excel_name_path)
    workbook.close()
    return worksheet

# 主函数


def export_to_excel(path, title, data, sheet_name='Default', data_type="count", template_path=None, template_sheet_name=None, spec_column=None):
    wb = open_file(path)
    ws = open_or_add_sheet(wb, sheet_name, template_path=template_path,
                           template_sheet_name=template_sheet_name)
    data_excel(path, wb, ws, title, data,
               data_type=data_type, spec_column=spec_column)
