# -*- encoding: utf-8 -*-
'''
@文件        :excel_util.py
@说明        :优化版本 - 提升Excel操作性能
@时间        :2022/09/15 10:26:29
@作者        :awx1192780
@版本        :1.1 - 性能优化版本
'''

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from util.common_tools import print_pro
from util.list_tools import one_layer
from datetime import datetime
from base import constants
from tqdm import tqdm
import pandas as pd
import json
import os
import copy
import time
from functools import lru_cache


__all__ = ['open_file', 'open_or_add_sheet', 'data_excel']


@lru_cache(maxsize=128)
def chrstep(c, num):
    """
    当前字母后num位的字母 - 缓存优化版本
    @param c: 字母
    @return chr
    """
    return chr(ord(c.upper())+num)


@lru_cache(maxsize=128)
def chr2index(c):
    """
    当前字母在字母表中顺序在'A'后的位数 - 缓存优化版本
    @param c: 字母
    @return int
    """
    return ord(c)-65


def smart_width_and_height(ws, _row, _list: list, spec_column: None):
    """
    智能调整单元格宽度和高度
    @param ws: 当前操作的worksheet
    @param _row: 当前操作的单元格给所在的行
    @param _list: 当前传入的数据
    @param spec_column: 需要特殊处理的列(字符比较长的列)
    @return 
    """
    for _index, _ in enumerate(_list):
        ws.column_dimensions[chr(_index+65).upper()].width = 14.67 if ws.column_dimensions[chr(
            _index+65).upper()].width < 14.67 else ws.column_dimensions[chr(_index+65).upper()].width
        if len(str(_)) > 30:
            ws.column_dimensions[chr(_index+65).upper()
                                 ].width = 22.5 + (len(str(_))/9) * 3

    # 根据content和desc中的内容调整单元格高度
    tar_height = 16 + 5 * (int(len(_list[chr2index(spec_column)])/7)+1) + \
        '\n'.join(map(str, _list)).count('\n') * 1

    ws.row_dimensions[_row].height = tar_height
    ws.row_dimensions[1].height = 30


def set_filter_and_sort(ws, _all_list: list):
    """
    设置过滤器和排序字段（有未解决bug，暂不可用）
    @param ws: 当前操作的worksheet
    @param _list: 当前传入的数据
    @return 
    """
    _row = len(_all_list)
    _column = len(_all_list[0])
    for _index, _ in enumerate(_all_list):
        ws.column_dimensions[chr(_index+65).upper()].width = 25
    if _index not in [0, 4]:
        ws.auto_filter.add_filter_column(_index, _all_list[_index])
    ws.auto_filter.ref = "A1:{}1".format(chr(_column+65-1).upper())
    ws.auto_filter.add_sort_condition("C2:D{}".format(_row))


def batch_insert(ws, _list, pos=[1, 1], spec_column=None, mode=0, _vertical='center'):
    """
    批量插入支持传入'A1' 和 [row,col]形式的数据 - 性能优化版本
    批量插入数据
    mode 0 横向 1 纵向
    @param ws: 当前操作的worksheet对象
    @param _list: 要插入的数据 list 格式
    @param pos: 起始单元格坐标 [row,column] 格式的 list 或 'A1' 格式的 str
    @param spec_column: 需要特殊处理的列(字符比较长的列)
    @param mode: 插入方向 0 水平 1 垂直
    @param _vertical: 设置当前单元格的文本格式
    @return 当前操作的worksheet对象
    """
    row, column = 1, 1
    if isinstance(pos, str):
        row, column = [chr2index(pos[:1])+1, int(pos[1:])]
    if isinstance(pos, list):
        row, column = pos
    
    # 预定义样式对象，避免重复创建
    alignment_left_top = Alignment(horizontal='left', vertical='top', wrapText=True)
    alignment_left_center = Alignment(horizontal='left', vertical='center', wrapText=True)
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")
    border_style = Border(top=double, left=thin, right=thin, bottom=double)
    bold_font = Font(bold=True)
    
    if mode == 0:
        # 批量设置单元格值，减少单个操作
        for index, value in enumerate(_list):
            cell = ws.cell(row=row, column=column+index, value=value)
            cell.alignment = alignment_left_top
            cell.border = border_style
            
            if row == 1:
                cell.font = bold_font
    else:
        for index, value in enumerate(_list):
            cell = ws.cell(row=row+index, column=column, value=value)
            cell.alignment = alignment_left_center
    
    return ws


def batch_insert_report(ws, _list, pos=[1, 1], spec_column=None, mode=0, _vertical='center', simple_mode=True):
    """
    批量插入报告数据 - 性能优化版本
    """
    row, column = 1, 1
    if isinstance(pos, str):
        row, column = [chr2index(pos[:1])+1, int(pos[1:])]
    if isinstance(pos, list):
        row, column = pos
    
    # 预定义样式对象，避免重复创建
    alignment_center = Alignment(horizontal='center', vertical='center', wrapText=not simple_mode)
    alignment_left_center = Alignment(horizontal='left', vertical='center', wrapText=not simple_mode)
    alignment_left_top = Alignment(horizontal='left', vertical='top', wrapText=not simple_mode)
    bold_font = Font(bold=True)
    
    if not simple_mode:
        smart_width_and_height(ws, row, _list, spec_column)
    if spec_column is not None and not simple_mode:
        ws.column_dimensions[spec_column].width = 85
        ws.column_dimensions[chrstep(spec_column, -1)].width = 35
    
    if mode == 0:
        for index, value in enumerate(_list):
            cell = ws.cell(row=row, column=column+index, value=value)
            cell.alignment = alignment_center
            
            if row == 1:
                cell.font = bold_font
            
            # 设置单元格为日期格式
            if spec_column and index == chr2index(spec_column)-5:
                cell.number_format = 'yyyy/m/d'
            
            # 以下是对 spec_column 列做特殊格式处理
            if spec_column is not None:
                # 根据《海外产出表2022》的格式调整，所有涉及的列改为 left,top
                if index in [99]:
                    cell.alignment = alignment_left_top
        
        if spec_column:
            ws.cell(row=1, column=chr2index(spec_column)+1).alignment = alignment_center
    else:
        for index, value in enumerate(_list):
            cell = ws.cell(row=row+index, column=column, value=value)
            cell.alignment = alignment_left_center
    
    return ws


def open_file(excel_name_path):
    """
    打开指定目录的excel
    @param excel_name_path: excel所在目录 
    @return Workbook
    """
    return Workbook() if not os.path.isfile(excel_name_path) else load_workbook(str(excel_name_path))


def open_or_add_sheet(workbook, sheet_name, template_path=None, template_sheet_name=None):
    """
    检查是否存在当前sheet_name, 若不存在则新建并以此命名第一个sheet
    @param workbook: 当前操作的workbook
    @param sheet_name: 工作簿名称
    @param template_path: 是否添加模板页面，若是填写其路径
    @param template_sheet_name: 模板页面工作簿名称
    @return Workbook
    """
    worksheet = workbook.active
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    elif len(workbook.sheetnames) != 0 and '月' in workbook.sheetnames[0]:
        worksheet = workbook.create_sheet(sheet_name)
    else:
        if template_path and template_sheet_name not in workbook.sheetnames:
            cp_sheet = copy_sheet_from(template_path, template_sheet_name)
            cp_sheet._parent = workbook
            workbook._add_sheet(cp_sheet)
        worksheet = workbook.create_sheet(sheet_name)
        if workbook.sheetnames[0] != template_sheet_name:
            del workbook[workbook.sheetnames[0]]
    return worksheet

# 执行插入数据


def copy_sheet_from(path, sheet_name):
    """
    从指定的excel中拷贝指定的工作簿
    @param path: 指定的excel路径
    @param sheet_name: 工作簿名称
    @return 指定的工作簿对应的worksheet
    """
    wb = open_file(path)
    if sheet_name in wb.sheetnames:
        worksheet = wb[sheet_name]
        worksheet.header = sheet_name
        return worksheet
    else:
        print_pro('不存在此工作簿!', constants.ERROR_PRINT)


def data_excel(excel_name_path, workbook, worksheet, header, data, data_type="count", template_path=None, template_sheet_name=None, spec_column=None, simple_mode=True):
    """
    最终处理，导出excel - 性能优化版本
    @param excel_name_path: 导出excel的目录
    @param workbook: 当前操作的workbook
    @param worksheet: 当前操作的worksheet
    @param header: header数据 list
    @param data: 源数据
    @param data_type: 处理方式: 'count'|'report' 分别表示统计各项指标数量 和 导出交付件详情列表
    @param template_path: 是否添加模板页面，若是填写其路径
    @param template_sheet_name: 模板页面工作簿名称
    @param spec_column: 需要特殊处理的列(字符比较长的列)
    @param simple_mode: 极简模式(可直接粘贴到在线交付件)|自动模式(自动格式化单元格)
    """
    if template_sheet_name and template_sheet_name not in workbook.sheetnames:
        cp_sheet = copy_sheet_from(template_path, template_sheet_name)
        cp_sheet._parent = workbook
        workbook._add_sheet(cp_sheet)
    
    # 优化数据处理，避免不必要的拷贝
    if isinstance(data, dict):
        values = list(data.values())
    else:
        values = data
    
    # 插入 header
    if data_type == 'count':
        worksheet = batch_insert(worksheet, header, [1, 1], spec_column)
    else:
        worksheet = batch_insert_report(worksheet, header, [1, 1], spec_column, simple_mode=simple_mode)
    
    # 批量插入数据，减少函数调用次数
    if data_type == 'count':
        for index, value in enumerate(values):
            worksheet = batch_insert(worksheet, value, [2+index, 1], spec_column)
    else:
        for index, value in enumerate(values):
            worksheet = batch_insert_report(worksheet, value, [2+index, 1], spec_column, simple_mode=simple_mode)

    # 暂无法使用filter和sort
    # set_filter_and_sort(worksheet, values)
    workbook.save(excel_name_path)
    workbook.close()


def to_count_excel(smart_dict,  template_path, template_sheet_name, infos=None):
    """
    统计月度各项指标，导出Excel
    最终处理，导出excel
    @param smart_dict: 源数据 dict 格式
    @param template_path: 是否添加模板页面，若是填写其路径
    @param template_sheet_name: 模板页面工作簿名称
    """
    try:
        alter_header = constants.HEADER
        alter_header.append('总计')
        alter_list = list(smart_dict.values())
        alter_list.append(sum(smart_dict.values()))
        df = pd.DataFrame([alter_list], columns=[alter_header])
        processed_data = df.values.tolist()
        if os.path.exists(constants.REPORT_FOLDER) == False:
            os.makedirs(constants.REPORT_FOLDER)
        wb = open_file(str(constants.REPORT_PATH))
        ws = open_or_add_sheet(wb, constants.SHEET_NAME, template_path=template_path,
                               template_sheet_name=template_sheet_name)
        report_path = str(constants.REPORT_PATH)
        if infos:
            report_path = report_path.replace('.xlsx',f'{infos}.xlsx')
            
        data_excel(str(constants.REPORT_PATH), wb, ws, alter_header, processed_data,
                   data_type='count', spec_column=None, simple_mode=True)

        print_pro('[processed_data] {} TO {}  => [{}]'.format(json.dumps(dict(zip(constants.HEADER, one_layer(
            processed_data))), indent=4, ensure_ascii=False), report_path, constants.SHEET_NAME), constants.SUCCESS_PRINT)

    except Exception as e:
        print_pro("导出失败，不存在目录或文件正在被使用。", constants.ERROR_PRINT, e)


def to_report_excel(report_dict, infos=None, template_path=None, template_sheet_name=None, spec_column=None, simple_mode=False):
    """
    统计月度交付件详情，导出Excel
    @param report_dict: 源数据 dict 格式
    @param template_path: 是否添加模板页面，若是填写其路径
    @param template_sheet_name: 模板页面工作簿名称
    @param spec_column: 需要特殊处理的列(字符比较长的列)
    @param simple_mode: 极简模式(可直接粘贴到在线交付件)|自动模式(自动格式化单元格)
    """
    try:
        df = pd.DataFrame(list(report_dict.values()),
                          columns=[constants.INFO_HEADER])
        processed_data = df.values.tolist()
        if os.path.exists(constants.REPORT_FOLDER) == False:
            os.makedirs(constants.REPORT_FOLDER)
        wb = open_file(str(constants.REPORT_PATH))
        report_path = str(constants.REPORT_PATH)
        if infos:
            report_path = report_path.replace('.xlsx',f'{infos}.xlsx')
            
        ws = open_or_add_sheet(wb, constants.REPORT_SHEET_NAME, template_path=template_path,
                               template_sheet_name=template_sheet_name)
        data_excel(report_path, wb, ws, constants.INFO_HEADER, processed_data,
                   data_type='report', spec_column=spec_column, simple_mode=simple_mode)

        print_pro('[report_data] {} TO {} => [{}]'.format(json.dumps(dict(zip(constants.INFO_HEADER, one_layer(
            processed_data))), indent=4, ensure_ascii=False), report_path, constants.REPORT_SHEET_NAME), p_type=constants.SUCCESS_PRINT)

    except Exception as e:
        print_pro("导出失败，不存在目录或文件正在被使用。", constants.ERROR_PRINT, e)
